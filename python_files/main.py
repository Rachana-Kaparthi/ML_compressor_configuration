import random
from convolution import *
from area_power import *
import itertools
# import pandas as pd
import openpyxl
from openpyxl import Workbook
import csv
import time
import multiprocessing
from functools import partial


def util(l, i, j, k, inputimage, kernal, exactop, count):
    start_time = time.time()

    # print(f"i = {i}, j = {j}, k = {k}, l = {l}")
    pos = [i, j, k, l]
    # pos = [1,4,5,6]
    print(pos)
    [approxop]=conv_3x3_sobel_approx(inputimage,kernal,pos)
    ssim = metrics.structural_similarity(exactop, approxop, data_range=1.0)
    mul = "mul" + str(count)

    # Hardware parameters extrapolation
    area=area_45nm(pos)
    power=power_45nm(pos)


    # # Write data to the Excel file
    # ws.cell(row=row_counter, column=1, value=mul)  # Multiplier
    # ws.cell(row=row_counter, column=2, value=str(pos))  # Position
    # ws.cell(row=row_counter, column=3, value=ssim)  # SSIM
    # ws.cell(row=row_counter, column=4, value=area)  # Area
    # ws.cell(row=row_counter, column=5, value=power)  # Power
    # ws.cell(row=row_counter, column=6, value=area * power)  # Cost function

    # Write data to the CSV file
    with open('output.csv', mode='a', newline='') as file:
        writer = csv.writer(file)
        writer.writerow([mul, pos, ssim, area, power, area * power])

    # row_counter += 1
    count += 1    
    print("            SSIM : ",ssim)
    print("            Cost Function :", area*power)
    print("Time taken ", time.time() - start_time)


if __name__ == "__main__":
    inputimage = imread('./ML_compressor_configuration/python_files/cameraman.png')
    inputimage = (inputimage/2).astype(int)

    # kernal1 = np.array([[-32, 0, 32],[-64,  0, 64],[-32, 0, 32]])
    # kernal2 = np.array([[-32, -64, -32],[0,  0, 0],[32, 64, 32]])

    kernal = np.array([[-127,-128,-89],[-31,10,52],[51,101,127]])


    [exactop]=conv_3x3_sobel_exact(inputimage,kernal)
    print("exacttop done")

    # Create a workbook and select the active sheet
    wb = Workbook()
    ws = wb.active

    # Write headers
    ws['A1'] = 'Multiplier'
    ws['B1'] = 'Position'
    ws['C1'] = 'SSIM'
    ws['D1'] = 'Area'
    ws['E1'] = 'Power'
    ws['F1'] = 'Cost function'

    # # Open a CSV file and create a CSV writer
    # csv_file = open('output.csv', mode='w', newline='')
    # csv_writer = csv.writer(csv_file)

    # # Write headers to CSV
    # csv_writer.writerow(['Multiplier', 'Position', 'SSIM', 'Area', 'Power', 'Cost function'])
    with open('output.csv', mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['Multiplier', 'Position', 'SSIM', 'Area', 'Power', 'Cost function'])


    # Generate all combinations of the ranges
    # combinations = itertools.product(range(1, 5), repeat=4)

    # Initialize row counter starting from the second row
    # row_counter = 2

    # # Create lists to store the combinations, SSIM values, and multipliers
    # data = []
    # SSIM = []
    # multiplier = []
    # AreaPowerProd = []
    # Initialize count
    count = 1


    # Print each combination
    for i in range(1,2):
        for j in range(1,2):
            for k in range(1,2):
                st = time.time()
                numbers = [1, 2, 3, 4, 5, 6, 7, 8]
                pool = multiprocessing.Pool(processes=8)
                partial_func = partial(util, i=i, j=j, k=k, inputimage=inputimage, kernal=kernal, exactop=exactop, count=count)
                pool.map(partial_func, numbers)
                pool.close()
                pool.join()
                print("Total Time taken ", time.time() - st)
                    

    # Save the workbook
    # wb.save('combinations.xlsx')


    # Close the CSV file
    # csv_file.close()

