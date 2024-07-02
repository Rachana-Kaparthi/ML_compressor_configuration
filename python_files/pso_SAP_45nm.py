import random
from convolution import *
from area_power import *
import itertools
# import pandas as pd
import openpyxl
from openpyxl import Workbook
import csv



inputimage = imread('cameraman.png')
inputimage = (inputimage/2).astype(int)

# kernal1 = np.array([[-32, 0, 32],[-64,  0, 64],[-32, 0, 32]])
# kernal2 = np.array([[-32, -64, -32],[0,  0, 0],[32, 64, 32]])

kernal = np.array([[-127,-128,-89],[-31,10,52],[51,101,127]])


[exactop]=conv_3x3_sobel_exact(inputimage,kernal)
print("exacttop done")

# Create a workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Write headers
ws['A1'] = 'Multiplier'
ws['B1'] = 'Position'
ws['C1'] = 'SSIM'
ws['D1'] = 'Area'
ws['E1'] = 'Power'
ws['F1'] = 'Cost function'

# Open a CSV file and create a CSV writer
csv_file = open('output.csv', mode='w', newline='')
csv_writer = csv.writer(csv_file)

# Write headers to CSV
csv_writer.writerow(['Multiplier', 'Position', 'SSIM', 'Area', 'Power', 'Cost function'])


# Generate all combinations of the ranges
combinations = itertools.product(range(1, 5), repeat=6)

# Initialize row counter starting from the second row
row_counter = 2

# # Create lists to store the combinations, SSIM values, and multipliers
# data = []
# SSIM = []
# multiplier = []
# AreaPowerProd = []
# Initialize count
count = 1


pos = [1,4,5,6,2,3]
print(pos)
[approxop]=conv_3x3_sobel_approx(inputimage,kernal,pos)
ssim = metrics.structural_similarity(exactop, approxop, data_range=1.0)
mul = "mul" + str(count)

# Hardware parameters extrapolation
area=area_45nm(pos)
power=power_45nm(pos)

# # appending the values to their respective list
# data.append(pos)
# SSIM.append(ssim)
# multiplier.append(mul)
# AreaPowerProd.append(area*power)

# Write data to the Excel file
ws.cell(row=row_counter, column=1, value=mul)  # Multiplier
ws.cell(row=row_counter, column=2, value=str(pos))  # Position
ws.cell(row=row_counter, column=3, value=ssim)  # SSIM
ws.cell(row=row_counter, column=4, value=area)  # Area
ws.cell(row=row_counter, column=5, value=power)  # Power
ws.cell(row=row_counter, column=6, value=area * power)  # Cost function

# Write data to the CSV file
csv_writer.writerow([mul, pos, ssim, area, power, area * power])


# # Print each combination
# for combo in combinations:
#     i, j, k, l, m, n = combo
#     # print(f"i = {i}, j = {j}, k = {k}, l = {l}")
#     pos = [i, j, k, l, m, n]
#     # pos = [1,4,5,6,2,3]
#     print(pos)
#     [approxop]=conv_3x3_sobel_approx(inputimage,kernal1,kernal2,pos)
#     ssim = metrics.structural_similarity(exactop, approxop, data_range=1.0)
#     mul = "mul" + str(count)

#     # Hardware parameters extrapolation
#     area=area_45nm(pos)
#     power=power_45nm(pos)

#     # # appending the values to their respective list
#     # data.append(pos)
#     # SSIM.append(ssim)
#     # multiplier.append(mul)
#     # AreaPowerProd.append(area*power)

#    # Write data to the Excel file
#     ws.cell(row=row_counter, column=1, value=mul)  # Multiplier
#     ws.cell(row=row_counter, column=2, value=str(pos))  # Position
#     ws.cell(row=row_counter, column=3, value=ssim)  # SSIM
#     ws.cell(row=row_counter, column=4, value=area)  # Area
#     ws.cell(row=row_counter, column=5, value=power)  # Power
#     ws.cell(row=row_counter, column=6, value=area * power)  # Cost function

#     # Write data to the CSV file
#     csv_writer.writerow([mul, pos, ssim, area, power, area * power])

#     row_counter += 1
#     count += 1    
#     print("            SSIM : ",ssim)

# Save the workbook
wb.save('combinations.xlsx')


# Close the CSV file
csv_file.close()

# # Convert the list to a DataFrame
# df = pd.DataFrame({
#     'Multiplier': multiplier,
#     'position': data,
#     'SSIM': SSIM,
#     'Cost function' : AreaPowerProd
# })

# # Write the DataFrame to an Excel file
# df.to_excel('combinations.xlsx', index=False)

# # Write the DataFrame to an CSV for dataset 
# df.to_csv('output.csv', index=False)


